from flask import abort, request, send_file
from flask_restful import Resource
from flask_jwt_extended import jwt_required
from openpyxl import load_workbook


from ysl.api import check_admin
from ysl.db import session
from ysl.db.interview import Interview
from ysl.db.evaluation import Evaluation
from ysl.db.interviewer import Interviewer
from ysl.db.interviewee import Interviewee
from ysl.db.question import Question
from ysl.db.access import Access


class CreateExcel(Resource):
    @jwt_required
    def get(self, agency_code, interview_id):
        check_admin(agency_code)

        workbook = load_workbook('./static/엔트리2차어드민.xlsx')

        interview = session.query(Interview).filter(Interview.interview_id == interview_id).first()

        worksheet = workbook.create_sheet()
        worksheet.title = interview.interview_name

        worksheet.cell(4, 1, '수험번호')
        worksheet.cell(4, 2, '이름')
        worksheet.cell(4, 3, '구분')

        interviewers = session.query(Access, Interviewer).join(Interviewer).filter(
                                     Access.interview == interview_id).all()
        questions = session.query(Question).filter(Question.interview == interview_id).all()
        interviewees = session.query(Interviewee).filter(Interviewee.interview == interview_id).all()

        for i in range(len(interviewers)):
            worksheet.cell(1, (i+1)*2+2, interviewers[i].Interviewer.email)
            worksheet.cell(2, (i+1)*2+2, interviewers[i].Interviewer.name)

        for i in range(len(questions)):
            worksheet.cell(3, i+4, questions[i].id)
            worksheet.cell(4, i+4, questions[i].title)

        for i in range(len(interviewees)):
            worksheet.cell(i+5, 1, interviewees[i].student_code)
            worksheet.cell(i+5, 2, interviewees[i].name)
            worksheet.cell(i+5, 3, interviewees[i].type)

            for x in range(len(interviewers)):
                interviewer = worksheet.cell(1, (x + 1) * 2 + 2).value
                for y in range(len(questions)):
                    question = worksheet.cell(3, 4+y).value

                    evaluation = (session.query(Evaluation).filter(Evaluation.interview == interview_id)
                                  .filter(Evaluation.interviewee == interviewees[i].student_code)
                                  .filter(Evaluation.interviewer == interviewer)
                                  .filter(Evaluation.question == question)).first()

                    answer = evaluation.answer if evaluation else "NULL"

                    worksheet.cell(i+5, 4+y, answer)

        workbook.save('./static/엔트리2차어드민.xlsx')

        return send_file('./static/엔트리2차어드민.xlsx', mimetype='file/xlsx')